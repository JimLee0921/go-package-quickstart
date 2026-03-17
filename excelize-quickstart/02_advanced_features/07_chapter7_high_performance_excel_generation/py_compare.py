import os
import time
import random
import xlsxwriter
from openpyxl import Workbook


# xlsxwriter 版本写入一百万条数据耗时：total cost: 65.8442s
def xlsxwriter_demo():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    files_dir = os.path.join(base_dir, "files")
    os.makedirs(files_dir, exist_ok=True)

    output_path = os.path.join(files_dir, "python_xlsxwriter.xlsx")

    start = time.time()

    workbook = xlsxwriter.Workbook(output_path, {"constant_memory": True})
    worksheet = workbook.add_worksheet("Sheet1")

    # 写表头
    worksheet.write_row(0, 0, ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"])
    worksheet.write_row(1, 0, ["start write data"])

    # 写数据（100万行）
    for row_idx in range(2, 1000000):
        row_data = [random.randint(0, 639999) for _ in range(10)]
        worksheet.write_row(row_idx, 0, row_data)

    workbook.close()

    total_cost = time.time() - start
    print(f"total cost: {total_cost:.4f}s")


# 83.8807s
def openpyxl_demo():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    files_dir = os.path.join(base_dir, "files")
    os.makedirs(files_dir, exist_ok=True)

    output_path = os.path.join(files_dir, "python_openpyxl.xlsx")

    start = time.time()

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")

    ws.append(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"])
    ws.append(["start write data"])

    for _ in range(2, 1000000):
        row_data = [random.randint(0, 639999) for _ in range(10)]
        ws.append(row_data)

    wb.save(output_path)

    total_cost = time.time() - start
    print(f"total cost: {total_cost:.4f}s")


if __name__ == "__main__":
    # xlsxwriter_demo()
    openpyxl_demo()